"""
Excel Template Generator for EarlyBird Import System
Aligns with emergent reference implementation format
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path

class EarlyBirdTemplateGenerator:
    """Generate Excel import templates"""
    
    def __init__(self):
        self.header_fill = PatternFill(
            start_color="1F4E78",  # Dark blue
            end_color="1F4E78",
            fill_type="solid"
        )
        self.header_font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.example_fill = PatternFill(
            start_color="E7E6E6",  # Light gray
            end_color="E7E6E6",
            fill_type="solid"
        )
    
    def _format_header(self, ws, num_cols):
        """Format header row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
    
    def _format_data_row(self, ws, row, num_cols):
        """Format data row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.example_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _set_column_widths(self, ws, widths):
        """Set column widths"""
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    
    def generate_customers_template(self, output_path: str):
        """Generate customers import template"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        ws_instructions.column_dimensions['A'].width = 80
        
        instructions = [
            ("CUSTOMER IMPORT TEMPLATE - EARLYBIRD", None),
            ("", None),
            ("INSTRUCTIONS:", None),
            ("1. Fill in the 'Customers' sheet with customer data", None),
            ("2. Ensure ALL required columns are filled:", None),
            ("   - Name (customer full name)", None),
            ("   - Email (valid email address)", None),
            ("   - Phone (10-digit phone number)", None),
            ("   - Area (delivery area/zone)", None),
            ("", None),
            ("OPTIONAL COLUMNS:", None),
            ("   - Subscription (subscription plan name)", None),
            ("   - Price (monthly subscription price)", None),
            ("   - Balance (account balance/credit)", None),
            ("", None),
            ("DATA FORMAT REQUIREMENTS:", None),
            ("   - Name: Text (max 100 characters)", None),
            ("   - Email: Valid email format (e.g., john@example.com)", None),
            ("   - Phone: 10 digits only (e.g., 9876543210)", None),
            ("   - Area: Text (max 50 characters)", None),
            ("   - Price: Number with 2 decimals (e.g., 500.00)", None),
            ("   - Balance: Number with 2 decimals (e.g., 1500.50)", None),
            ("", None),
            ("DO NOT:", None),
            ("   - Delete or modify header row", None),
            ("   - Add extra columns", None),
            ("   - Leave required fields empty", None),
            ("   - Use special characters in phone numbers", None),
            ("", None),
            ("EXAMPLE:", None),
            ("   See 'Customers' sheet for sample data", None),
        ]
        
        for row, (text, _) in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1).value = text
            ws_instructions.cell(row=row, column=1).font = Font(size=11)
        
        # Create Customers sheet
        ws_customers = wb.create_sheet("Customers", 1)
        
        headers = ["Name", "Email", "Phone", "Area", "Subscription", "Price", "Balance"]
        for col, header in enumerate(headers, 1):
            ws_customers.cell(row=1, column=col).value = header
        
        self._format_header(ws_customers, len(headers))
        
        # Add sample data
        sample_data = [
            ["Rajesh Kumar", "rajesh@email.com", "9876543210", "Bangalore", "Premium Daily", 500.00, 1500.00],
            ["Priya Sharma", "priya@email.com", "9876543211", "Hyderabad", "Standard Weekly", 300.00, 800.00],
            ["Amit Patel", "amit@email.com", "9876543212", "Pune", "Premium Daily", 500.00, 0.00],
            ["Sunita Singh", "sunita@email.com", "9876543213", "Delhi", "Basic Monthly", 200.00, 500.00],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_customers.cell(row=row_idx, column=col_idx).value = value
            self._format_data_row(ws_customers, row_idx, len(headers))
        
        # Add 10 empty rows for data entry
        for row_idx in range(len(sample_data) + 2, len(sample_data) + 12):
            for col_idx in range(1, len(headers) + 1):
                ws_customers.cell(row=row_idx, column=col_idx).border = self.border
        
        self._set_column_widths(ws_customers, [20, 25, 15, 20, 20, 12, 12])
        
        # Freeze first row
        ws_customers.freeze_panes = "A2"
        
        wb.save(output_path)
        print(f"✅ Customer template created: {output_path}")
    
    def generate_orders_template(self, output_path: str):
        """Generate orders import template"""
        wb = openpyxl.Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Instructions sheet
        ws_instr = wb.create_sheet("Instructions", 0)
        ws_instr.column_dimensions['A'].width = 80
        
        instructions = [
            "ORDER IMPORT TEMPLATE - EARLYBIRD",
            "",
            "REQUIRED COLUMNS:",
            "   - Customer ID (customer identifier)",
            "   - Order Date (YYYY-MM-DD format)",
            "   - Amount (order total amount)",
            "",
            "OPTIONAL COLUMNS:",
            "   - Delivery Date (YYYY-MM-DD format)",
            "   - Status (pending/delivered/cancelled)",
            "",
            "DATA FORMAT:",
            "   - Customer ID: CUST001, CUST002, etc.",
            "   - Order Date: 2026-01-23",
            "   - Delivery Date: 2026-01-25",
            "   - Amount: Number with 2 decimals",
            "   - Status: pending, out_for_delivery, delivered, cancelled",
        ]
        
        for row, text in enumerate(instructions, 1):
            ws_instr.cell(row=row, column=1).value = text
        
        # Orders sheet
        ws_orders = wb.create_sheet("Orders", 1)
        headers = ["Customer ID", "Order Date", "Amount", "Delivery Date", "Status"]
        
        for col, header in enumerate(headers, 1):
            ws_orders.cell(row=1, column=col).value = header
        
        self._format_header(ws_orders, len(headers))
        
        # Sample data
        today = datetime.now()
        sample_data = [
            ["CUST001", today.strftime("%Y-%m-%d"), 500.00, (today + timedelta(days=1)).strftime("%Y-%m-%d"), "pending"],
            ["CUST002", today.strftime("%Y-%m-%d"), 300.00, (today + timedelta(days=1)).strftime("%Y-%m-%d"), "pending"],
            ["CUST003", (today - timedelta(days=1)).strftime("%Y-%m-%d"), 500.00, today.strftime("%Y-%m-%d"), "delivered"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_orders.cell(row=row_idx, column=col_idx).value = value
            self._format_data_row(ws_orders, row_idx, len(headers))
        
        # Empty rows
        for row_idx in range(len(sample_data) + 2, len(sample_data) + 12):
            for col_idx in range(1, len(headers) + 1):
                ws_orders.cell(row=row_idx, column=col_idx).border = self.border
        
        self._set_column_widths(ws_orders, [15, 15, 12, 15, 20])
        ws_orders.freeze_panes = "A2"
        
        wb.save(output_path)
        print(f"✅ Order template created: {output_path}")
    
    def generate_delivery_template(self, output_path: str):
        """Generate delivery import template"""
        wb = openpyxl.Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Instructions
        ws_instr = wb.create_sheet("Instructions", 0)
        ws_instr.column_dimensions['A'].width = 80
        
        instructions = [
            "DELIVERY IMPORT TEMPLATE - EARLYBIRD",
            "",
            "REQUIRED COLUMNS:",
            "   - Customer ID",
            "   - Customer Name",
            "   - Delivery Date (YYYY-MM-DD)",
            "   - Area (delivery zone/location)",
            "",
            "OPTIONAL COLUMNS:",
            "   - Address (detailed delivery address)",
            "   - Status (pending/delivered/failed)",
            "",
            "DATE FORMAT: YYYY-MM-DD (e.g., 2026-01-23)",
        ]
        
        for row, text in enumerate(instructions, 1):
            ws_instr.cell(row=row, column=1).value = text
        
        # Delivery sheet
        ws_delivery = wb.create_sheet("Delivery", 1)
        headers = ["Customer ID", "Customer Name", "Delivery Date", "Area", "Address", "Status"]
        
        for col, header in enumerate(headers, 1):
            ws_delivery.cell(row=1, column=col).value = header
        
        self._format_header(ws_delivery, len(headers))
        
        # Sample data
        today = datetime.now()
        sample_data = [
            ["CUST001", "Rajesh Kumar", today.strftime("%Y-%m-%d"), "Bangalore", "123 Main St, Bangalore", "pending"],
            ["CUST002", "Priya Sharma", today.strftime("%Y-%m-%d"), "Hyderabad", "456 Oak Ave, Hyderabad", "pending"],
            ["CUST003", (today - timedelta(days=1)).strftime("%Y-%m-%d"), "Pune", "789 Pine Rd, Pune", "delivered"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_delivery.cell(row=row_idx, column=col_idx).value = value
            self._format_data_row(ws_delivery, row_idx, len(headers))
        
        for row_idx in range(len(sample_data) + 2, len(sample_data) + 12):
            for col_idx in range(1, len(headers) + 1):
                ws_delivery.cell(row=row_idx, column=col_idx).border = self.border
        
        self._set_column_widths(ws_delivery, [15, 20, 15, 20, 30, 15])
        ws_delivery.freeze_panes = "A2"
        
        wb.save(output_path)
        print(f"✅ Delivery template created: {output_path}")
    
    def generate_all_templates(self, output_dir: str = "."):
        """Generate all import templates"""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        print("📊 Generating EarlyBird Import Templates...")
        print("")
        
        self.generate_customers_template(str(output_dir / "EarlyBird_Customer_Import_Template.xlsx"))
        self.generate_orders_template(str(output_dir / "EarlyBird_Orders_Import_Template.xlsx"))
        self.generate_delivery_template(str(output_dir / "EarlyBird_Delivery_Import_Template.xlsx"))
        
        print("")
        print("✅ All templates generated successfully!")
        print(f"📁 Output directory: {output_dir}")

if __name__ == "__main__":
    # Generate templates
    generator = EarlyBirdTemplateGenerator()
    
    # Output to assets folder
    output_path = Path("src/frontend/assets")
    generator.generate_all_templates(str(output_path))
