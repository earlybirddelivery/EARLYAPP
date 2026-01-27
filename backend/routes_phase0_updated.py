from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timedelta, date as date_type
import uuid
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models_phase0_updated import (
    Customer, CustomerCreate, CustomerUpdate, CustomerConfirm,
    Subscription, SubscriptionCreate, SubscriptionUpdate,
    DeliveryBoy, DeliveryBoyCreate,
    Product, ProductCreate,
    DeliveryListItem, DashboardStats,
    CustomerImportRow, ImportPreview, ImportResult,
    CustomerStatus, SubscriptionMode, SubscriptionStatus
)
from database import db
from auth import get_current_user, hash_password

router = APIRouter(prefix="/phase0-v2", tags=["Phase 0 Updated"])

# ==================== PRODUCT ROUTES ====================

@router.post("/products", response_model=Product)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    """Create a new product"""
    product_doc = {
        "id": str(uuid.uuid4()),
        **product.model_dump()
    }
    await db.products.insert_one(product_doc)
    return product_doc

@router.get("/products", response_model=List[Product])
async def get_products(current_user: dict = Depends(get_current_user)):
    """Get all products"""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    return products

# ==================== IMAGE UPLOAD ====================

@router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload house image and return URL"""
    import base64
    
    # Read file content
    contents = await file.read()
    
    # Convert to base64 for simple storage (in production, use S3/cloud storage)
    encoded = base64.b64encode(contents).decode('utf-8')
    image_url = f"data:image/{file.filename.split('.')[-1]};base64,{encoded}"
    
    return {"image_url": image_url}

# ==================== CUSTOMER ROUTES ====================

@router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    """Create a new customer (default status: trial)"""
    customer_doc = {
        "id": str(uuid.uuid4()),
        **customer.model_dump()
    }
    
    # STEP 21: Create linked user record if not already provided
    if not customer.user_id:
        # Generate unique email for the customer
        user_email = f"customer-{customer_doc['id']}@earlybird.local"
        
        # Check if user with this email already exists (shouldn't happen but be safe)
        existing_user = await db.users.find_one({"email": user_email}, {"_id": 0})
        if not existing_user:
            # Default password for new customer users
            default_password = "earlybird2025"
            
            # Create user record linked to customer
            user_doc = {
                "id": str(uuid.uuid4()),
                "email": user_email,
                "name": customer.name,
                "phone": customer.phone,
                "role": "customer",
                "customer_v2_id": customer_doc["id"],  # STEP 21: Link to customer
                "password_hash": hash_password(default_password),
                "is_active": True,
                "created_at": datetime.utcnow().isoformat()
            }
            result = await db.users.insert_one(user_doc)
            
            # Link user back to customer
            customer_doc["user_id"] = user_doc["id"]
            print(f"[STEP 21] Created linked user for customer {customer_doc['id']}: {user_email}")
    
    await db.customers_v2.insert_one(customer_doc)
    return customer_doc

# Combined endpoint kept for backward compatibility but subscription is optional
@router.post("/customers-with-subscription")
async def create_customer_with_subscription(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Create customer with optional subscription"""
    customer_data = data.get("customer", {})
    subscription_data = data.get("subscription")
    
    # Create customer
    customer_doc = {
        "id": str(uuid.uuid4()),
        **customer_data
    }
    
    # STEP 21: Create linked user record if not already provided
    if not customer_data.get("user_id"):
        user_email = f"customer-{customer_doc['id']}@earlybird.local"
        existing_user = await db.users.find_one({"email": user_email}, {"_id": 0})
        if not existing_user:
            default_password = "earlybird2025"
            user_doc = {
                "id": str(uuid.uuid4()),
                "email": user_email,
                "name": customer_data.get("name"),
                "phone": customer_data.get("phone"),
                "role": "customer",
                "customer_v2_id": customer_doc["id"],
                "password_hash": hash_password(default_password),
                "is_active": True,
                "created_at": datetime.utcnow().isoformat()
            }
            await db.users.insert_one(user_doc)
            customer_doc["user_id"] = user_doc["id"]
            print(f"[STEP 21] Created linked user for customer {customer_doc['id']}: {user_email}")
    
    await db.customers_v2.insert_one(customer_doc)
    
    result = {
        "customer": customer_doc,
        "message": "Customer created successfully"
    }
    
    # Create subscription only if provided
    if subscription_data:
        subscription_doc = {
            "id": str(uuid.uuid4()),
            "customer_id": customer_doc["id"],
            **subscription_data
        }
        await db.subscriptions_v2.insert_one(subscription_doc)
        result["subscription"] = subscription_doc
        result["message"] = "Customer and subscription created successfully"
    
    return result

@router.get("/customers", response_model=List[Customer])
async def get_customers(
    status: Optional[str] = None,
    area: Optional[str] = None,
    marketing_boy: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all customers with advanced filtering and role-based access"""
    query = {}
    if status:
        query["status"] = status
    if area:
        query["area"] = area
    if marketing_boy:
        query["marketing_boy"] = marketing_boy
    
    # Role-based filtering: Marketing staff see only their customers
    if current_user.get("role") == "marketing_staff":
        marketing_name = current_user.get("name") or current_user.get("email") or ""
        query["marketing_boy_id"] = current_user.get("id")
    
    customers = await db.customers_v2.find(query, {"_id": 0}).to_list(1000)
    return customers

@router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific customer"""
    customer = await db.customers_v2.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(
    customer_id: str,
    updates: CustomerUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a customer"""
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    result = await db.customers_v2.update_one({"id": customer_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    customer = await db.customers_v2.find_one({"id": customer_id}, {"_id": 0})
    return customer

@router.patch("/customers/{customer_id}")
async def patch_customer(
    customer_id: str,
    updates: dict,
    current_user: dict = Depends(get_current_user)
):
    """Partially update a customer (accepts any fields)"""
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    # Log the update for debugging
    print(f"PATCH customer {customer_id} with updates: {updates}")
    
    result = await db.customers_v2.update_one({"id": customer_id}, {"$set": updates})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Verify the update
    customer = await db.customers_v2.find_one({"id": customer_id}, {"_id": 0})
    print(f"Customer after update: delivery_boy_id={customer.get('delivery_boy_id')}, shift={customer.get('shift')}")
    
    return {"message": "Customer updated successfully", "customer": customer}

@router.post("/customers/confirm")
async def confirm_customer(
    confirm: CustomerConfirm,
    current_user: dict = Depends(get_current_user)
):
    """Confirm a trial customer to active status"""
    result = await db.customers_v2.update_one(
        {"id": confirm.customer_id, "status": "trial"},
        {"$set": {"status": "active"}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found or already confirmed")
    
    return {"message": "Customer confirmed successfully", "customer_id": confirm.customer_id}

@router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a customer"""
    result = await db.customers_v2.delete_one({"id": customer_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Also delete related subscriptions
    await db.subscriptions_v2.delete_many({"customer_id": customer_id})
    
    return {"message": "Customer deleted successfully"}

# ==================== SUBSCRIPTION ROUTES ====================

@router.post("/subscriptions", response_model=Subscription)
async def create_subscription(
    subscription: SubscriptionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new subscription with full validation"""
    from subscription_engine_v2 import subscription_engine
    
    # Verify customer exists
    customer = await db.customers_v2.find_one({"id": subscription.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Verify product exists if provided
    if subscription.product_id:
        product = await db.products.find_one({"id": subscription.product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
    
    # Validate subscription data
    subscription_dict = subscription.model_dump()
    is_valid, error = subscription_engine.validate_subscription(subscription_dict)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error)
    
    subscription_doc = {
        "id": str(uuid.uuid4()),
        **subscription_dict,
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    
    await db.subscriptions_v2.insert_one(subscription_doc)
    
    # Log audit entry
    await db.subscription_audit.insert_one({
        "subscription_id": subscription_doc["id"],
        "user_id": current_user.get("id"),
        "action": "created",
        "details": f"Created {subscription_dict.get('mode')} subscription",
        "timestamp": datetime.now().isoformat()
    })
    
    return subscription_doc

@router.get("/subscriptions", response_model=List[Subscription])
async def get_subscriptions(
    customer_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all subscriptions, optionally filtered by customer"""
    query = {"customer_id": customer_id} if customer_id else {}
    subscriptions = await db.subscriptions_v2.find(query, {"_id": 0}).to_list(1000)
    return subscriptions

@router.get("/subscriptions/{subscription_id}", response_model=Subscription)
async def get_subscription(subscription_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific subscription"""
    subscription = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    return subscription

@router.put("/subscriptions/{subscription_id}", response_model=Subscription)
async def update_subscription(
    subscription_id: str,
    updates: SubscriptionUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a subscription with audit logging"""
    from subscription_engine_v2 import subscription_engine
    
    # Get existing subscription
    existing = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    # Validate product if being updated
    if "product_id" in update_data and update_data["product_id"]:
        product = await db.products.find_one({"id": update_data["product_id"]}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=400, detail="Product not found")
    
    # Add updated timestamp
    update_data["updated_at"] = datetime.now().isoformat()
    
    result = await db.subscriptions_v2.update_one({"id": subscription_id}, {"$set": update_data})
    
    # Log audit entries for each field changed
    for field, new_value in update_data.items():
        if field != "updated_at" and existing.get(field) != new_value:
            await db.subscription_audit.insert_one({
                "subscription_id": subscription_id,
                "user_id": current_user.get("id"),
                "action": f"updated_{field}",
                "old_value": existing.get(field),
                "new_value": new_value,
                "timestamp": datetime.now().isoformat()
            })
    
    subscription = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    return subscription

@router.delete("/subscriptions/{subscription_id}")
async def delete_subscription(subscription_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a subscription"""
    result = await db.subscriptions_v2.delete_one({"id": subscription_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    return {"message": "Subscription deleted successfully"}

# ==================== DELIVERY BOY ROUTES ====================

@router.post("/delivery-boys", response_model=DeliveryBoy)
async def create_delivery_boy(
    delivery_boy: DeliveryBoyCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new delivery boy and corresponding user account"""
    delivery_boy_id = str(uuid.uuid4())
    delivery_boy_doc = {
        "id": delivery_boy_id,
        **delivery_boy.model_dump()
    }
    await db.delivery_boys_v2.insert_one(delivery_boy_doc)
    
    # Create user account for the delivery boy
    # Generate email from name
    email = f"{delivery_boy.name.lower().replace(' ', '_')}@earlybird.com"
    
    # Check if email already exists
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    if existing_user:
        # Add a unique suffix if email exists
        email = f"{delivery_boy.name.lower().replace(' ', '')}{delivery_boy_id[:4]}@earlybird.com"
    
    # Create user with the same ID as delivery boy
    default_password = "delivery123"
    user_doc = {
        "id": delivery_boy_id,
        "email": email,
        "phone": None,
        "name": delivery_boy.name,
        "password_hash": hash_password(default_password),
        "role": "delivery_boy",
        "is_active": True
    }
    await db.users.insert_one(user_doc)
    
    return delivery_boy_doc

@router.get("/delivery-boys", response_model=List[DeliveryBoy])
async def get_delivery_boys(
    area: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all delivery boys, optionally filtered by area"""
    query = {"area_assigned": area} if area else {}
    delivery_boys = await db.delivery_boys_v2.find(query, {"_id": 0}).to_list(1000)
    return delivery_boys

# ==================== HELPER: CALCULATE QUANTITY ====================

def calculate_quantity_for_date(subscription: dict, target_date: str) -> float:
    """
    Calculate delivery quantity using subscription engine
    """
    from subscription_engine_v2 import subscription_engine
    return subscription_engine.compute_qty(target_date, subscription)

# ==================== DELIVERY LIST GENERATION ====================

@router.get("/delivery/generate", response_model=List[DeliveryListItem])
async def generate_delivery_list(
    date: str,
    area: Optional[str] = None,
    delivery_boy_id: Optional[str] = None,
    shift: Optional[str] = None,  # morning, evening, both, or all
    current_user: dict = Depends(get_current_user)
):
    """
    Generate delivery list for a specific date with priority-based quantity calculation
    Includes: Active customers + Trial customers whose trial_start_date <= date
    """
    from subscription_engine_v2 import subscription_engine
    from datetime import datetime
    
    # Get ACTIVE customers + TRIAL customers whose start date has passed
    query = {
        "$or": [
            {"status": "active"},
            {
                "status": "trial",
                "$or": [
                    {"trial_start_date": {"$lte": date}},
                    {"trial_start_date": None},
                    {"trial_start_date": {"$exists": False}}
                ]
            }
        ]
    }
    customers = await db.customers_v2.find(query, {"_id": 0}).to_list(1000)
    
    # Get only ACTIVE subscriptions with auto_start=True (draft subscriptions excluded)
    subscriptions = await db.subscriptions_v2.find({
        "status": "active",
        "auto_start": True
    }, {"_id": 0}).to_list(1000)
    
    # Get all products
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    product_map = {p["id"]: p for p in products}
    
    # If filtering by delivery_boy, get their area
    if delivery_boy_id:
        delivery_boy = await db.delivery_boys_v2.find_one({"id": delivery_boy_id}, {"_id": 0})
        if delivery_boy:
            area = delivery_boy["area_assigned"]
    
    # Build customer lookup
    customer_map = {c["id"]: c for c in customers}
    
    # Build subscription lookup by customer_id
    subscription_map = {}
    for sub in subscriptions:
        customer_id = sub.get("customerId") or sub.get("customer_id")
        if customer_id not in subscription_map:
            subscription_map[customer_id] = []
        subscription_map[customer_id].append(sub)
    
    delivery_list = []
    serial = 1
    
    for customer in customers:
        # Filter by area if specified
        if area and customer.get("area") != area:
            continue
        
        # Get customer's subscriptions
        customer_subs = subscription_map.get(customer["id"], [])
        
        for subscription in customer_subs:
            # Calculate quantity for the date
            quantity = calculate_quantity_for_date(subscription, date)
            
            if quantity <= 0:
                continue
            
            # Get shift for this date
            sub_shift = subscription.get("shift", "morning")
            shift_overrides = subscription.get("shift_overrides", [])
            for override in shift_overrides:
                if override.get("date") == date:
                    sub_shift = override.get("shift")
                    break
            
            # Filter by shift if specified
            if shift and shift != "all":
                if shift != sub_shift:
                    continue
            
            # Get product info
            product_id = subscription.get("productId") or subscription.get("product_id")
            product = product_map.get(product_id, {})
            
            delivery_list.append({
                "serial": serial,
                "customer_id": customer["id"],
                "customer_name": customer["name"],
                "phone": customer["phone"],
                "address": customer["address"],
                "area": customer.get("area", ""),
                "product_id": product_id,
                "product_name": product.get("name", "Unknown"),
                "quantity": quantity,
                "shift": sub_shift,
                "price_per_unit": subscription.get("price_per_unit", 0),
                "notes": customer.get("notes", ""),
                "status": customer.get("status", "active").capitalize(),
                "map_link": customer.get("map_link", ""),
                "subscription_id": subscription.get("id"),
                "delivery_boy_id": customer.get("delivery_boy_id"),
                "delivery_boy_name": customer.get("delivery_boy_name")
            })
            serial += 1
    
    return delivery_list

@router.get("/delivery/whatsapp-format")
async def get_whatsapp_format(
    date: str,
    area: Optional[str] = None,
    delivery_boy_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate WhatsApp formatted text for delivery list with added products"""
    delivery_list = await generate_delivery_list(date, area, delivery_boy_id, current_user)
    
    # Get delivery boy name if filtered
    delivery_boy_name = None
    if delivery_boy_id:
        delivery_boy = await db.delivery_boys_v2.find_one({"id": delivery_boy_id}, {"_id": 0})
        if delivery_boy:
            delivery_boy_name = delivery_boy["name"]
    
    # Get added products for this date
    added_products = await db.added_products_v2.find({
        "date": date
    }, {"_id": 0}).to_list(1000)
    
    # Filter added products by area if specified
    if area:
        added_products = [p for p in added_products if p.get("area") == area]
    
    # Header with delivery boy name if filtered
    if delivery_boy_name:
        lines = [f"📅 Delivery List for {date} - {delivery_boy_name}\n"]
    else:
        lines = [f"📅 Delivery List for {date}\n"]
    
    lines.append(f"*Total: {len(delivery_list)} customers")
    if added_products:
        lines.append(f" + {len(added_products)} added products*")
    else:
        lines.append("*")
    
    lines.append('─' * 40)
    lines.append("")
    
    # Create table format
    lines.append("```")
    lines.append(f"# | Customer      | Phone      | Product        | Qty | Area")
    lines.append('-' * 70)
    
    rowNum = 1
    for item in delivery_list:
        customer_name = str(item['customer_name'])[:13].ljust(13)
        phone = str(item['phone'])[:10].ljust(10)
        product_name = str(item['product_name'])[:14].ljust(14)
        qty = str(int(item['quantity'])).ljust(3)
        area_name = str(item['area'])[:15]
        
        lines.append(f"{str(rowNum).ljust(2)} | {customer_name} | {phone} | {product_name} | {qty} | {area_name}")
        rowNum += 1
    
    # Add section for added products
    if added_products:
        lines.append('-' * 70)
        lines.append("")
        lines.append("➕ ADDITIONAL PRODUCTS")
        lines.append('-' * 70)
        
        for product in added_products:
            customer_name = str(product.get('customer_name', 'N/A'))[:13].ljust(13)
            phone = str(product.get('phone', 'N/A'))[:10].ljust(10)
            product_name = str(product.get('product_name', 'N/A'))[:14].ljust(14)
            qty = str(int(product.get('quantity', 0))).ljust(3)
            area_name = str(product.get('area', 'N/A'))[:15]
            
            lines.append(f"{str(rowNum).ljust(2)} | {customer_name} | {phone} | {product_name} | {qty} | {area_name}")
            rowNum += 1
    
    # Add totals by area
    area_totals = {}
    for item in delivery_list:
        area_name = item['area']
        if area_name not in area_totals:
            area_totals[area_name] = 0
        area_totals[area_name] += item['quantity']
    
    # Add added products to totals
    for product in added_products:
        area_name = product.get('area')
        if area_name not in area_totals:
            area_totals[area_name] = 0
        area_totals[area_name] += product.get('quantity', 0)
    
    lines.append('-' * 70)
    lines.append("📊 Totals by Area:")
    for area_name, total in area_totals.items():
        lines.append(f"   {area_name}: {int(total)}L")
    lines.append("```")
    
    return {"text": "\n".join(lines)}

# ==================== EXCEL EXPORT (DATE RANGE) ====================

@router.get("/delivery/export-excel")
async def export_excel(
    start_date: str,
    end_date: str,
    area: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Export delivery list as Excel for date range
    Format: CustomerName | Phone | Area | Product | Price | Status | date1 | date2 | ... | Total Qty | Total Bill
    """
    from datetime import datetime as dt, timedelta
    
    start_dt = dt.strptime(start_date, "%Y-%m-%d").date()
    end_dt = dt.strptime(end_date, "%Y-%m-%d").date()
    
    # Generate list of dates
    date_list = []
    current = start_dt
    while current <= end_dt:
        date_list.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    
    # Get customers and subscriptions
    query = {"status": {"$in": ["active", "trial"]}}
    if area:
        query["area"] = area
    
    customers = await db.customers_v2.find(query, {"_id": 0}).to_list(1000)
    subscriptions = await db.subscriptions_v2.find({}, {"_id": 0}).to_list(1000)
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    product_map = {p["id"]: p for p in products}
    subscription_map = {}
    for sub in subscriptions:
        customer_id = sub.get("customerId") or sub.get("customer_id")
        if customer_id not in subscription_map:
            subscription_map[customer_id] = []
        subscription_map[customer_id].append(sub)
    
    # Build data rows
    rows = []
    for customer in customers:
        customer_subs = subscription_map.get(customer["id"], [])
        
        # Group by product
        product_data = {}
        for sub in customer_subs:
            product_id = sub["product_id"]
            if product_id not in product_data:
                product_data[product_id] = {
                    "product_name": product_map.get(product_id, {}).get("name", "Unknown"),
                    "price_per_unit": sub.get("price_per_unit", 0),
                    "quantities": {}
                }
            
            # Calculate quantities for each date
            for date_str in date_list:
                qty = calculate_quantity_for_date(sub, date_str)
                if date_str not in product_data[product_id]["quantities"]:
                    product_data[product_id]["quantities"][date_str] = 0
                product_data[product_id]["quantities"][date_str] += qty
        
        # Create one row per customer (aggregated by product)
        for product_id, data in product_data.items():
            row = {
                "CustomerName": customer["name"],
                "Phone": customer["phone"],
                "Area": customer.get("area", ""),
                "Product": data["product_name"],
                "Price": data["price_per_unit"],
                "Status": customer.get("status", "active").capitalize()
            }
            
            # Add date columns
            total_qty = 0
            for date_str in date_list:
                qty = data["quantities"].get(date_str, 0)
                row[date_str] = qty
                total_qty += qty
            
            row["Total Qty"] = total_qty
            row["Total Bill"] = total_qty * data["price_per_unit"]
            
            rows.append(row)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Report"
    
    # Header
    headers = ["CustomerName", "Phone", "Area", "Product", "Price", "Status"] + date_list + ["Total Qty", "Total Bill"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=delivery_report_{start_date}_to_{end_date}.xlsx"}
    )

# ==================== EXCEL IMPORT ====================

@router.post("/customers/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview Excel/CSV import with validation"""
    try:
        contents = await file.read()
        
        # Read file
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
        
        # Validate columns - only name, phone, area are mandatory
        required_columns = ['name', 'phone', 'area']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        errors = []
        valid_rows = 0
        preview_data = []
        
        for idx, row in df.iterrows():
            row_errors = []
            
            # Validate required fields - only name, phone, area
            if pd.isna(row['name']) or not str(row['name']).strip():
                row_errors.append("Name is required")
            if pd.isna(row['phone']) or not str(row['phone']).strip():
                row_errors.append("Phone is required")
            if pd.isna(row.get('area')) or not str(row.get('area', '')).strip():
                row_errors.append("Area is required")
            
            if row_errors:
                errors.append({
                    "row": idx + 2,  # Excel row number (1-indexed + header)
                    "errors": row_errors
                })
            else:
                valid_rows += 1
            
            # Add to preview (first 10 rows)
            if len(preview_data) < 10:
                preview_data.append({
                    "row": idx + 2,
                    "name": str(row['name']),
                    "phone": str(row['phone']),
                    "area": str(row.get('area', '')),
                    "default_daily_qty": float(row.get('default_daily_qty', 0)) if not pd.isna(row.get('default_daily_qty')) else 0,
                    "errors": row_errors
                })
        
        return {
            "total_rows": len(df),
            "valid_rows": valid_rows,
            "errors": errors[:20],  # First 20 errors
            "preview_data": preview_data
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@router.post("/customers/import")
async def import_customers(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import customers from comprehensive Excel/CSV with all details"""
    try:
        contents = await file.read()
        
        # Read file
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
        
        imported_count = 0
        failed_count = 0
        errors = []
        
        # Get all products for lookup
        products = await db.products.find({}, {"_id": 0}).to_list(100)
        product_map = {p["name"].lower(): p for p in products}
        
        # Get all delivery boys
        delivery_boys = await db.delivery_boys_v2.find({}, {"_id": 0}).to_list(100)
        delivery_boy_map = {db["name"].lower(): db for db in delivery_boys}
        
        # Get all marketing users
        marketing_users = await db.users.find({"role": "marketing_staff"}, {"_id": 0}).to_list(100)
        marketing_user_map = {u["name"].lower(): u for u in marketing_users}
        
        for idx, row in df.iterrows():
            try:
                # Validate required fields
                if pd.isna(row.get('name')) or pd.isna(row.get('phone')) or pd.isna(row.get('area')):
                    failed_count += 1
                    customer_name = row.get('name', 'Unknown') if not pd.isna(row.get('name')) else 'Unknown'
                    errors.append({
                        "row": idx + 2, 
                        "customer_name": customer_name,
                        "error": "Missing required fields: name, phone, or area",
                        "error_type": "ValidationError"
                    })
                    continue
                
                # Get or create delivery boy
                delivery_boy_id = None
                if not pd.isna(row.get('delivery_boy')) and str(row['delivery_boy']).strip():
                    delivery_boy_name = str(row['delivery_boy']).strip()
                    delivery_boy_lower = delivery_boy_name.lower()
                    
                    if delivery_boy_lower in delivery_boy_map:
                        delivery_boy_id = delivery_boy_map[delivery_boy_lower]["id"]
                    else:
                        # Create new delivery boy
                        new_boy_id = str(uuid.uuid4())
                        new_boy = {
                            "id": new_boy_id,
                            "name": delivery_boy_name,
                            "area_assigned": str(row['area'])
                        }
                        await db.delivery_boys_v2.insert_one(new_boy)
                        delivery_boy_id = new_boy_id
                        delivery_boy_map[delivery_boy_lower] = new_boy
                        
                        # Create user account for delivery boy
                        from passlib.context import CryptContext
                        pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
                        
                        # Check if user already exists
                        email = f"{delivery_boy_lower.replace(' ', '_')}@earlybird.com"
                        existing_user = await db.users.find_one({"email": email}, {"_id": 0})
                        
                        if not existing_user:
                            default_password = "delivery123"  # Default password
                            user_doc = {
                                "id": new_boy_id,
                                "email": email,
                                "phone": None,
                                "name": delivery_boy_name,
                                "password_hash": pwd_context.hash(default_password),
                                "role": "delivery_boy",
                                "is_active": True
                            }
                            await db.users.insert_one(user_doc)
                            print(f"[OK] Created delivery boy user: {email} (password: {default_password})")
                
                # Get or create marketing boy
                marketing_boy = None
                marketing_boy_id = None
                if not pd.isna(row.get('marketing_boy')) and str(row['marketing_boy']).strip():
                    marketing_boy_name = str(row['marketing_boy']).strip()
                    marketing_boy_lower = marketing_boy_name.lower()
                    
                    if marketing_boy_lower in marketing_user_map:
                        marketing_boy = marketing_user_map[marketing_boy_lower]["name"]
                        marketing_boy_id = marketing_user_map[marketing_boy_lower]["id"]
                    else:
                        marketing_boy = marketing_boy_name
                
                # Parse status
                status = str(row.get('status', 'trial')).strip().lower()
                if status not in ['trial', 'active', 'paused', 'stopped']:
                    status = 'trial'
                
                # Parse shift
                shift = str(row.get('shift', 'morning')).strip().lower()
                if shift not in ['morning', 'evening', 'both']:
                    shift = 'morning'
                
                # Parse trial start date
                trial_start_date = None
                if not pd.isna(row.get('trial_start_date')):
                    try:
                        trial_start_date = str(row['trial_start_date'])[:10]  # YYYY-MM-DD
                    except:
                        pass
                
                # Parse previous balance
                previous_balance = 0
                if not pd.isna(row.get('previous_balance')):
                    try:
                        previous_balance = float(row['previous_balance'])
                    except:
                        previous_balance = 0
                
                # Create customer
                customer_doc = {
                    "id": str(uuid.uuid4()),
                    "name": str(row['name']).strip(),
                    "phone": str(row['phone']).strip(),
                    "address": str(row.get('address', '')) if not pd.isna(row.get('address')) else '',
                    "area": str(row['area']).strip(),
                    "map_link": str(row.get('map_link', '')) if not pd.isna(row.get('map_link')) else None,
                    "location": None,
                    "status": status,
                    "trial_start_date": trial_start_date,
                    "notes": str(row.get('notes', '')) if not pd.isna(row.get('notes')) else None,
                    "house_image_url": None,
                    "marketing_boy": marketing_boy,
                    "marketing_boy_id": marketing_boy_id,
                    "delivery_boy_id": delivery_boy_id,
                    "previous_balance": previous_balance,
                    "custom_product_prices": {}
                }
                
                await db.customers_v2.insert_one(customer_doc)
                
                # Create user account for customer
                from passlib.context import CryptContext
                pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
                
                # Generate email from phone number
                customer_phone = str(row['phone']).strip().replace('+', '').replace('-', '').replace(' ', '')
                customer_email = f"{customer_phone}@customer.earlybird.com"
                
                # Check if user already exists
                existing_customer_user = await db.users.find_one({"email": customer_email}, {"_id": 0})
                
                if not existing_customer_user:
                    default_password = "customer123"  # Default password
                    customer_user_doc = {
                        "id": customer_doc["id"],
                        "email": customer_email,
                        "phone": customer_phone,
                        "name": customer_doc["name"],
                        "password_hash": pwd_context.hash(default_password),
                        "role": "customer",
                        "is_active": True
                    }
                    await db.users.insert_one(customer_user_doc)
                    print(f"[OK] Created customer user: {customer_email} (password: {default_password})")
                
                # Create subscription if product info is provided
                if not pd.isna(row.get('product_name')) and str(row.get('product_name', '')).strip():
                    product_name = str(row['product_name']).strip().lower()
                    print(f"DEBUG: Creating subscription for {customer_doc['name']}, product: {product_name}")
                    print(f"DEBUG: Available products: {list(product_map.keys())}")
                    
                    if product_name in product_map:
                        product = product_map[product_name]
                        
                        # Get price (custom or default)
                        price = product["default_price"]
                        if not pd.isna(row.get('product_price')):
                            try:
                                price = float(row['product_price'])
                            except:
                                pass
                        
                        # Get quantity (in packets for milk, direct for water tin)
                        default_qty = 0
                        if not pd.isna(row.get('default_qty_packets')):
                            try:
                                packets = float(row['default_qty_packets'])
                                # Convert packets to liters for milk products
                                if product["unit"] == "Liter" or product["unit"] == "L":
                                    default_qty = packets * 0.5  # 1 packet = 0.5L
                                else:
                                    default_qty = packets  # Direct for water tin
                            except:
                                pass
                        
                        # Parse subscription type
                        sub_type = str(row.get('subscription_type', 'fixed_daily')).strip().lower()
                        mode = "fixed_daily"
                        weekly_pattern = None
                        
                        if sub_type.startswith('weekly_pattern'):
                            mode = "weekly_pattern"
                            # Parse pattern like "weekly_pattern:0,2,4"
                            if ':' in sub_type:
                                pattern_str = sub_type.split(':')[1]
                                try:
                                    weekly_pattern = [int(d.strip()) for d in pattern_str.split(',')]
                                except:
                                    weekly_pattern = [0, 2, 4]  # Default Mon, Wed, Fri
                        
                        # Parse daily quantities from day_1 to day_31 columns
                        day_overrides = []
                        import calendar
                        from datetime import datetime
                        
                        # Get current month/year
                        now = datetime.now()
                        year = now.year
                        month = now.month
                        max_days = calendar.monthrange(year, month)[1]
                        
                        print(f"DEBUG: Parsing daily data for {customer_doc['name']}, max_days={max_days}")
                        
                        for day in range(1, 32):
                            # Skip invalid days for the month
                            if day > max_days:
                                continue
                                
                            col_name = f"day_{day}"
                            if col_name in row and not pd.isna(row.get(col_name)):
                                try:
                                    packets = float(row[col_name])
                                    if packets > 0:
                                        # Convert to liters if milk product
                                        if product["unit"] == "Liter" or product["unit"] == "L":
                                            actual_qty = packets * 0.5
                                        else:
                                            actual_qty = packets
                                        
                                        date_str = f"{year}-{month:02d}-{day:02d}"
                                        
                                        day_overrides.append({
                                            "date": date_str,
                                            "quantity": actual_qty,
                                            "shift": shift
                                        })
                                except Exception as day_err:
                                    # Log the error but continue processing
                                    print(f"Warning: Failed to parse day_{day} for customer {row.get('name', 'Unknown')}: {str(day_err)}")
                                    pass
                        
                        subscription_doc = {
                            "id": str(uuid.uuid4()),
                            "customer_id": customer_doc["id"],
                            "product_id": product["id"],
                            "price_per_unit": price,
                            "mode": mode,
                            "default_qty": default_qty,
                            "shift": shift,
                            "weekly_pattern": weekly_pattern,
                            "day_overrides": day_overrides,
                            "irregular_list": [],
                            "shift_overrides": [],
                            "pause_intervals": [],
                            "stop_date": None,
                            "status": "active" if status == "active" else "draft",
                            "auto_start": True if status == "active" else False
                        }
                        await db.subscriptions_v2.insert_one(subscription_doc)
                
                imported_count += 1
            
            except Exception as e:
                failed_count += 1
                customer_name = row.get('name', 'Unknown') if not pd.isna(row.get('name')) else 'Unknown'
                errors.append({
                    "row": idx + 2,
                    "customer_name": customer_name,
                    "error": str(e),
                    "error_type": type(e).__name__
                })
        
        return {
            "success": True,
            "imported_count": imported_count,
            "failed_count": failed_count,
            "total_rows": len(df),
            "errors": errors  # Return ALL errors, not just first 20
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error importing file: {str(e)}")

# ==================== DASHBOARD ====================

@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(
    stat_date: str,
    current_user: dict = Depends(get_current_user)
):
    """Get dashboard statistics with role-based filtering"""
    # Get customers based on role
    query = {}
    if current_user.get("role") == "marketing_staff":
        # Marketing staff see only their customers
        query["marketing_boy_id"] = current_user.get("id")
    
    all_customers = await db.customers_v2.find(query, {"_id": 0}).to_list(1000)
    trial_customers = [c for c in all_customers if c.get("status") == "trial"]
    active_customers = [c for c in all_customers if c.get("status") == "active"]
    
    # Get customer IDs for subscription filtering
    customer_ids = [c["id"] for c in all_customers]
    
    # Get subscriptions (filter by customer if marketing staff)
    if current_user.get("role") == "marketing_staff":
        subscriptions = await db.subscriptions_v2.find({"customer_id": {"$in": customer_ids}}, {"_id": 0}).to_list(1000)
    else:
        subscriptions = await db.subscriptions_v2.find({}, {"_id": 0}).to_list(1000)
    
    # Get delivery boys for name mapping
    delivery_boys = await db.delivery_boys_v2.find({}, {"_id": 0}).to_list(1000)
    delivery_boy_map = {db["id"]: db["name"] for db in delivery_boys}
    
    # Calculate liters by area, delivery boy, and marketing boy
    customer_map = {c["id"]: c for c in all_customers}
    liters_by_area = {}
    liters_by_delivery_boy = {}
    liters_by_marketing_boy = {}
    
    for subscription in subscriptions:
        # Handle both camelCase and snake_case for backward compatibility
        customer_id = subscription.get("customerId") or subscription.get("customer_id")
        customer = customer_map.get(customer_id)
        if not customer:
            continue
        
        qty = calculate_quantity_for_date(subscription, stat_date)
        
        # By area
        area = customer.get("area", "Unknown")
        if area not in liters_by_area:
            liters_by_area[area] = 0
        liters_by_area[area] += qty
        
        # By delivery boy
        delivery_boy_id = customer.get("delivery_boy_id") or ""
        delivery_boy_name = delivery_boy_map.get(delivery_boy_id, "Unassigned")
        if delivery_boy_name is None or delivery_boy_name == "":
            delivery_boy_name = "Unassigned"
        if delivery_boy_name not in liters_by_delivery_boy:
            liters_by_delivery_boy[delivery_boy_name] = 0
        liters_by_delivery_boy[delivery_boy_name] += qty
        
        # By marketing boy
        marketing_boy = customer.get("marketing_boy") or "Unassigned"
        if marketing_boy is None:
            marketing_boy = "Unassigned"
        if marketing_boy not in liters_by_marketing_boy:
            liters_by_marketing_boy[marketing_boy] = 0
        liters_by_marketing_boy[marketing_boy] += qty
    
    # Find top performers
    top_delivery_boy = None
    if liters_by_delivery_boy:
        top_boy = max(liters_by_delivery_boy.items(), key=lambda x: x[1])
        top_delivery_boy = {"name": top_boy[0], "liters": top_boy[1]}
    
    top_marketing_boy = None
    if liters_by_marketing_boy:
        top_marketing = max(liters_by_marketing_boy.items(), key=lambda x: x[1])
        top_marketing_boy = {"name": top_marketing[0], "liters": top_marketing[1]}
    
    return {
        "total_customers": len(all_customers),
        "trial_customers": len(trial_customers),
        "active_customers": len(active_customers),
        "total_subscriptions": len(subscriptions),
        "liters_by_area": liters_by_area,
        "liters_by_delivery_boy": liters_by_delivery_boy,
        "liters_by_marketing_boy": liters_by_marketing_boy,
        "top_delivery_boy": top_delivery_boy,
        "top_marketing_boy": top_marketing_boy
    }

@router.get("/areas")
async def get_all_areas(current_user: dict = Depends(get_current_user)):
    """Get all areas with main/sub area hierarchy"""
    areas = await db.areas_v2.find({}, {"_id": 0}).to_list(1000)
    
    # Filter for new format areas only (with full_name field)
    structured_areas = [a for a in areas if "full_name" in a]
    
    # If no structured areas, get from customers (backward compatibility)
    if not structured_areas:
        customers = await db.customers_v2.find({}, {"_id": 0, "area": 1}).to_list(1000)
        unique_areas = sorted(list(set([c.get("area") for c in customers if c.get("area")])))
        return {"areas": unique_areas, "structured_areas": []}
    
    return {"structured_areas": structured_areas, "areas": [a["full_name"] for a in structured_areas]}

@router.get("/customer/{customer_id}/calendar")
async def get_customer_calendar(
    customer_id: str,
    month: str,  # YYYY-MM format
    product_id: Optional[str] = None,  # Filter by specific product or "all"
    current_user: dict = Depends(get_current_user)
):
    """Get customer's delivery calendar for a specific month with multi-product support"""
    from datetime import datetime
    import calendar
    
    # Parse month
    year, month_num = map(int, month.split('-'))
    
    # Get customer and their subscriptions
    customer = await db.customers_v2.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get subscriptions - filter by product if specified
    query = {"customer_id": customer_id}
    if product_id and product_id != "all":
        query["product_id"] = product_id
    
    subscriptions = await db.subscriptions_v2.find(query, {"_id": 0}).to_list(100)
    
    # Get all products for this customer
    all_products = await db.products.find({}, {"_id": 0}).to_list(100)
    product_map = {p["id"]: p for p in all_products}
    
    # Generate calendar data for the month
    num_days = calendar.monthrange(year, month_num)[1]
    calendar_data = []
    
    for day in range(1, num_days + 1):
        date_str = f"{year}-{month_num:02d}-{day:02d}"
        day_data = {
            "date": date_str,
            "day_name": datetime(year, month_num, day).strftime("%A"),
            "products": []  # Changed from "deliveries" to "products" for clarity
        }
        
        # Calculate deliveries for this date across all subscriptions
        for subscription in subscriptions:
            from subscription_engine_v2 import subscription_engine
            qty = subscription_engine.compute_qty(date_str, subscription)
            
            product_id = subscription.get("productId") or subscription.get("product_id")
            product = product_map.get(product_id)
            product_name = product.get("name", "Unknown") if product else "Unknown"
            product_unit = product.get("unit", "L") if product else "L"
            
            # Get shift for this date (check overrides first, then default)
            shift = subscription.get("shift", "morning")
            shift_overrides = subscription.get("shift_overrides", [])
            for override in shift_overrides:
                if override.get("date") == date_str:
                    shift = override.get("shift")
                    break
            
            day_data["products"].append({
                "subscription_id": subscription["id"],
                "product_id": product_id,
                "product_name": product_name,
                "product_unit": product_unit,
                "quantity": qty,
                "shift": shift,
                "status": subscription.get("status", "draft"),
                "is_main_subscription": subscription.get("mode") in ["fixed_daily", "weekly_pattern", "day_by_day"]
            })
        
        calendar_data.append(day_data)
    
    return {
        "customer": customer,
        "month": month,
        "calendar": calendar_data,
        "available_products": [{"id": p["id"], "name": p["name"], "unit": p["unit"]} for p in all_products]
    }

@router.post("/customer/{customer_id}/pause")
async def pause_customer_delivery(
    customer_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Pause deliveries for a customer for a specific period"""
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    subscription_id = data.get("subscription_id")
    
    if not start_date:
        raise HTTPException(status_code=400, detail="start_date is required")
    
    # Get subscription
    subscription = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    # Add pause interval
    pause_intervals = subscription.get("pause_intervals", [])
    pause_intervals.append({
        "start": start_date,
        "end": end_date
    })
    
    await db.subscriptions_v2.update_one(
        {"id": subscription_id},
        {"$set": {"pause_intervals": pause_intervals}}
    )
    
    return {"message": "Pause added successfully", "pause_intervals": pause_intervals}

@router.post("/customer/{customer_id}/override")
async def set_quantity_override(
    customer_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Set quantity override for specific date"""
    date_str = data.get("date")
    quantity = data.get("quantity")
    subscription_id = data.get("subscription_id")
    
    if not date_str or quantity is None:
        raise HTTPException(status_code=400, detail="date and quantity are required")
    
    subscription = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    # Add or update day override
    day_overrides = subscription.get("day_overrides", [])
    
    # Remove existing override for this date if exists
    day_overrides = [d for d in day_overrides if d.get("date") != date_str]
    
    # Add new override
    day_overrides.append({
        "date": date_str,
        "quantity": float(quantity)
    })
    
    await db.subscriptions_v2.update_one(
        {"id": subscription_id},
        {"$set": {"day_overrides": day_overrides}}
    )
    
    return {"message": "Quantity override set successfully"}

@router.post("/customer/{customer_id}/add-product")
async def add_product_to_date(
    customer_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add a new product to customer for specific date(s) - creates irregular subscription"""
    product_id = data.get("product_id")
    date_str = data.get("date")
    quantity = data.get("quantity")
    shift = data.get("shift", "morning")  # Default to morning if not specified
    
    if not product_id or not date_str or quantity is None:
        raise HTTPException(status_code=400, detail="product_id, date, and quantity are required")
    
    # Check if subscription already exists for this product
    existing_sub = await db.subscriptions_v2.find_one({
        "customer_id": customer_id,
        "product_id": product_id
    }, {"_id": 0})
    
    if existing_sub:
        # Add to irregular list
        irregular_list = existing_sub.get("irregular_list", [])
        
        # Remove existing entry for this date if exists
        irregular_list = [item for item in irregular_list if item.get("date") != date_str]
        
        # Add new irregular entry
        irregular_list.append({
            "date": date_str,
            "quantity": float(quantity),
            "shift": shift
        })
        
        await db.subscriptions_v2.update_one(
            {"id": existing_sub["id"]},
            {"$set": {"irregular_list": irregular_list}}
        )
    else:
        # Create new irregular-only subscription
        product = await db.products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        new_subscription = {
            "id": str(uuid.uuid4()),
            "customer_id": customer_id,
            "product_id": product_id,
            "price_per_unit": product.get("default_price", 0),
            "mode": "irregular",
            "default_qty": 0,
            "shift": shift,
            "weekly_pattern": None,
            "day_overrides": [],
            "irregular_list": [{
                "date": date_str,
                "quantity": float(quantity),
                "shift": shift
            }],
            "shift_overrides": [],
            "pause_intervals": [],
            "stop_date": None,
            "status": "active",
            "auto_start": True
        }
        
        await db.subscriptions_v2.insert_one(new_subscription)
    
    return {"message": "Product added successfully to delivery date"}

@router.post("/customer/{customer_id}/change-shift")
async def change_delivery_shift(
    customer_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Change delivery shift for specific date"""
    subscription_id = data.get("subscription_id")
    date_str = data.get("date")
    new_shift = data.get("shift")  # morning, evening, or both
    
    if not subscription_id or not date_str or not new_shift:
        raise HTTPException(status_code=400, detail="subscription_id, date, and shift are required")
    
    subscription = await db.subscriptions_v2.find_one({"id": subscription_id}, {"_id": 0})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    
    # Add or update shift override
    shift_overrides = subscription.get("shift_overrides", [])
    
    # Remove existing override for this date
    shift_overrides = [s for s in shift_overrides if s.get("date") != date_str]
    
    # Add new shift override
    shift_overrides.append({
        "date": date_str,
        "shift": new_shift
    })
    
    await db.subscriptions_v2.update_one(
        {"id": subscription_id},
        {"$set": {"shift_overrides": shift_overrides}}
    )
    
    return {"message": f"Shift changed to {new_shift} for {date_str}"}

@router.get("/download-sample-template")
async def download_sample_template(current_user: dict = Depends(get_current_user)):
    """Download comprehensive Excel template for customer import with daily delivery columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Import Template"
    
    # Headers - comprehensive format with dates 1-31
    headers = [
        "name", "phone", "address", "area", "delivery_boy", "marketing_boy", 
        "shift", "status", "trial_start_date", "product_name", "product_price", 
        "subscription_type", "default_qty_packets", "previous_balance", "notes", "map_link"
    ]
    
    # Add date columns 1-31
    for day in range(1, 32):
        headers.append(f"day_{day}")
    
    ws.append(headers)
    
    # Sample data with all fields including daily deliveries
    sample_data = [
        # Full example with subscription and daily data
        ["Rajesh Kumar", "9876543210", "123 Main St, Koramangala", "Koramangala", "Amit Singh", "Marketing Staff",
         "morning", "active", "", "Full Cream Milk", "60", "fixed_daily", "4", "0", "Gate code: 1234", "https://goo.gl/maps/xyz"] 
         + [4] * 15 + [0] * 16,  # 4 packets for first 15 days, 0 for rest
        
        # Trial customer with start date and partial deliveries
        ["Priya Sharma", "9876543211", "456 Park Ave, Indiranagar", "Indiranagar", "Rajesh Kumar", "Marketing Staff",
         "evening", "trial", "2025-12-15", "Toned Milk", "55", "fixed_daily", "3", "150", "Call before delivery", ""]
         + [0] * 10 + [2] * 21,  # No delivery first 10 days, then 2 packets
        
        # Customer with weekly pattern and varying quantities
        ["Vikram Rao", "9876543212", "789 Oak Rd, Whitefield", "Whitefield", "Suresh Patel", "",
         "both", "active", "", "Full Cream Milk", "60", "weekly_pattern:0,2,4", "2", "0", "", ""]
         + [2, 0, 2, 0, 2, 0, 0] * 4 + [2, 0, 2],  # Mon/Wed/Fri pattern
        
        # Water Tin customer with occasional deliveries
        ["Deepa Nair", "9876543213", "101 Palm Gardens, HSR Layout", "HSR Layout", "Amit Singh", "Marketing Staff",
         "morning", "active", "", "20L Water Tin", "50", "fixed_daily", "2", "0", "", ""]
         + [1] * 20 + [0] * 11,  # 1 tin for 20 days, none after
        
        # Minimal - only required fields, no daily data
        ["Minimal Customer", "9999999999", "Some Address", "Bommanahalli", "", "", "morning", "trial", "", "", "", "", "", "0", "", ""]
         + [""] * 31  # Empty daily columns
    ]
    
    # Add sample data first
    for row in sample_data:
        ws.append(row)
    
    # Add instruction rows at the bottom
    ws.append([])
    ws.append(["INSTRUCTIONS:"])
    ws.append(["MANDATORY FIELDS: name, phone, address, area"])
    ws.append(["shift: morning, evening, or both"])
    ws.append(["status: trial or active"])
    ws.append(["subscription_type: fixed_daily or weekly_pattern:0,2,4 (0=Mon, 6=Sun)"])
    ws.append(["default_qty_packets: For milk products, enter packets (1 packet = 0.5L). For Water Tin, enter tins."])
    ws.append(["previous_balance: Any previous balance amount to carry forward"])
    ws.append(["day_1 to day_31: Enter delivered quantities in packets for each day (leave blank or 0 for no delivery)"])
    ws.append(["If delivery_boy or marketing_boy names don't exist, they will be created automatically"])
    
    for row in sample_data:
        ws.append(row)
    
    # Style header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_import_template.xlsx"}
    )

# ==================== USER MANAGEMENT FOR DROPDOWNS ====================

@router.get("/users")
async def get_users_by_role(role: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get users filtered by role (for dropdowns)"""
    query = {}
    if role:
        query["role"] = role
    
    users = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    return users

# ==================== AREAS MANAGEMENT ====================

@router.post("/areas")
async def create_area(data: dict, current_user: dict = Depends(get_current_user)):
    """Create new area with main and sub area"""
    main_area = data.get("main_area")
    sub_area = data.get("sub_area")
    delivery_boy_ids = data.get("delivery_boy_ids", [])  # Multiple delivery boys
    
    if not main_area or not sub_area:
        raise HTTPException(status_code=400, detail="main_area and sub_area are required")
    
    area_full_name = f"{main_area} - {sub_area}"
    
    # Check if area already exists
    existing = await db.areas_v2.find_one({"full_name": area_full_name})
    if existing:
        raise HTTPException(status_code=400, detail="Area already exists")
    
    # Create area document
    area_doc = {
        "id": str(uuid.uuid4()),
        "main_area": main_area,
        "sub_area": sub_area,
        "full_name": area_full_name,
        "delivery_boy_ids": delivery_boy_ids
    }
    
    await db.areas_v2.insert_one(area_doc)
    
    # Add area to each delivery boy's assigned areas
    for boy_id in delivery_boy_ids:
        boy = await db.delivery_boys_v2.find_one({"id": boy_id})
        if boy:
            assigned_areas = boy.get("assigned_areas", [])
            if area_full_name not in assigned_areas:
                assigned_areas.append(area_full_name)
            await db.delivery_boys_v2.update_one(
                {"id": boy_id},
                {"$set": {"assigned_areas": assigned_areas}}
            )
    
    return {"message": "Area created successfully", "area": area_doc}

@router.put("/areas/{area_id}")
async def update_area(area_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update area details"""
    main_area = data.get("main_area")
    sub_area = data.get("sub_area")
    delivery_boy_ids = data.get("delivery_boy_ids", [])
    
    if not main_area or not sub_area:
        raise HTTPException(status_code=400, detail="main_area and sub_area are required")
    
    area_full_name = f"{main_area} - {sub_area}"
    
    # Get old area
    old_area = await db.areas_v2.find_one({"id": area_id}, {"_id": 0})
    if not old_area:
        raise HTTPException(status_code=404, detail="Area not found")
    
    old_full_name = old_area["full_name"]
    
    # Update area document
    await db.areas_v2.update_one(
        {"id": area_id},
        {"$set": {
            "main_area": main_area,
            "sub_area": sub_area,
            "full_name": area_full_name,
            "delivery_boy_ids": delivery_boy_ids
        }}
    )
    
    # Remove old area from all delivery boys
    await db.delivery_boys_v2.update_many(
        {},
        {"$pull": {"assigned_areas": old_full_name}}
    )
    
    # Add new area to selected delivery boys
    for boy_id in delivery_boy_ids:
        boy = await db.delivery_boys_v2.find_one({"id": boy_id})
        if boy:
            assigned_areas = boy.get("assigned_areas", [])
            if area_full_name not in assigned_areas:
                assigned_areas.append(area_full_name)
            await db.delivery_boys_v2.update_one(
                {"id": boy_id},
                {"$set": {"assigned_areas": assigned_areas}}
            )
    
    # Update customers using old area name
    await db.customers_v2.update_many(
        {"area": old_full_name},
        {"$set": {"area": area_full_name}}
    )
    
    return {"message": "Area updated successfully"}

@router.delete("/areas/{area_id}")
async def delete_area(area_id: str, current_user: dict = Depends(get_current_user)):
    """Delete area (doesn't delete customers)"""
    area = await db.areas_v2.find_one({"id": area_id}, {"_id": 0})
    if not area:
        raise HTTPException(status_code=404, detail="Area not found")
    
    area_name = area["full_name"]
    
    # Remove from areas collection
    await db.areas_v2.delete_one({"id": area_id})
    
    # Remove from all delivery boys
    await db.delivery_boys_v2.update_many(
        {},
        {"$pull": {"assigned_areas": area_name}}
    )
    
    return {"message": "Area deleted successfully"}

@router.put("/delivery-boy/{boy_id}/assign-area")
async def assign_area_to_delivery_boy(
    boy_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Assign area to delivery boy"""
    area = data.get("area")
    
    if not area:
        raise HTTPException(status_code=400, detail="area is required")
    
    result = await db.delivery_boys_v2.update_one(
        {"id": boy_id},
        {"$set": {"area_assigned": area}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Delivery boy not found")
    
    return {"message": "Area assigned successfully"}

