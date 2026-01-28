"""
analytics_engine.py - Core analytics engine for EarlyBird
Aggregates data from multiple collections for comprehensive business insights
"""

from datetime import datetime, timedelta, date
from bson import ObjectId
from typing import Dict, List, Any, Optional
from database import db
import json
from decimal import Decimal


class AnalyticsEngine:
    """Core analytics engine for business metrics and reporting"""

    # ==================== REVENUE ANALYTICS ====================

    @staticmethod
    async def get_revenue_overview(start_date: Optional[str] = None, 
                                   end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Get revenue overview with daily breakdown
        
        Returns:
        {
            "total_revenue": float,
            "daily_revenue": [{"date": "2024-01-15", "amount": 5000}],
            "daily_orders": [{"date": "2024-01-15", "count": 25}],
            "average_order_value": float,
            "top_products": [{"product": "Milk", "revenue": 15000}],
            "payment_methods": [{"method": "UPI", "revenue": 10000}],
            "period": {"start": "2024-01-01", "end": "2024-01-31"}
        }
        """
        try:
            # Parse dates (default: last 30 days)
            if not end_date:
                end_date = datetime.now()
            else:
                end_date = datetime.fromisoformat(end_date)
            
            if not start_date:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = datetime.fromisoformat(start_date)

            # Get all orders in date range
            orders = await db.orders.find({
                "created_at": {
                    "$gte": start_date,
                    "$lte": end_date
                },
                "status": {"$ne": "CANCELLED"}
            }).to_list(None)

            # Calculate totals
            total_revenue = sum(float(order.get("total_amount", 0)) for order in orders)
            total_orders = len(orders)
            avg_order_value = total_revenue / total_orders if total_orders > 0 else 0

            # Daily breakdown
            daily_data = {}
            for order in orders:
                date_key = order.get("created_at", datetime.now()).strftime("%Y-%m-%d")
                if date_key not in daily_data:
                    daily_data[date_key] = {"revenue": 0, "orders": 0}
                daily_data[date_key]["revenue"] += float(order.get("total_amount", 0))
                daily_data[date_key]["orders"] += 1

            daily_revenue = [
                {"date": date_key, "amount": daily_data[date_key]["revenue"]}
                for date_key in sorted(daily_data.keys())
            ]

            daily_orders = [
                {"date": date_key, "count": daily_data[date_key]["orders"]}
                for date_key in sorted(daily_data.keys())
            ]

            # Top products
            product_revenue = {}
            for order in orders:
                items = order.get("items", [])
                for item in items:
                    product_name = item.get("product_name", "Unknown")
                    quantity = item.get("quantity", 0)
                    price = float(item.get("price", 0))
                    revenue = quantity * price

                    if product_name not in product_revenue:
                        product_revenue[product_name] = 0
                    product_revenue[product_name] += revenue

            top_products = [
                {"product": product, "revenue": revenue}
                for product, revenue in sorted(
                    product_revenue.items(),
                    key=lambda x: x[1],
                    reverse=True
                )[:10]
            ]

            # Payment methods
            payment_methods_data = {}
            for order in orders:
                method = order.get("payment_method", "Unknown")
                if method not in payment_methods_data:
                    payment_methods_data[method] = 0
                payment_methods_data[method] += float(order.get("total_amount", 0))

            payment_methods = [
                {"method": method, "revenue": amount}
                for method, amount in sorted(
                    payment_methods_data.items(),
                    key=lambda x: x[1],
                    reverse=True
                )
            ]

            return {
                "total_revenue": round(total_revenue, 2),
                "total_orders": total_orders,
                "average_order_value": round(avg_order_value, 2),
                "daily_revenue": daily_revenue,
                "daily_orders": daily_orders,
                "top_products": top_products,
                "payment_methods": payment_methods,
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat()
                }
            }

        except Exception as e:
            raise Exception(f"Revenue analytics error: {str(e)}")

    # ==================== CUSTOMER ANALYTICS ====================

    @staticmethod
    async def get_customer_metrics(start_date: Optional[str] = None,
                                   end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Get customer insights and metrics
        
        Returns:
        {
            "total_customers": int,
            "new_customers": int,
            "repeat_customers": int,
            "customer_retention": float (percentage),
            "customer_ltv": float,
            "top_customers": [{"customer_id": "...", "spending": 5000}],
            "customer_segments": [{"segment": "HIGH_VALUE", "count": 50}],
            "average_order_frequency": float
        }
        """
        try:
            if not end_date:
                end_date = datetime.now()
            else:
                end_date = datetime.fromisoformat(end_date)

            if not start_date:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = datetime.fromisoformat(start_date)

            # Total unique customers
            all_customers = await db.customers_v2.find().to_list(None)
            total_customers = len(all_customers)

            # New customers in period
            new_customers_list = await db.customers_v2.find({
                "created_at": {
                    "$gte": start_date,
                    "$lte": end_date
                }
            }).to_list(None)
            new_customers = len(new_customers_list)

            # Customer orders
            orders = await db.orders.find({
                "created_at": {
                    "$gte": start_date,
                    "$lte": end_date
                }
            }).to_list(None)

            # Customer spending analysis
            customer_spending = {}
            customer_order_count = {}
            
            for order in orders:
                customer_id = order.get("customer_id")
                if not customer_id:
                    continue

                if customer_id not in customer_spending:
                    customer_spending[customer_id] = 0
                    customer_order_count[customer_id] = 0

                customer_spending[customer_id] += float(order.get("total_amount", 0))
                customer_order_count[customer_id] += 1

            # Repeat customers (more than 1 order)
            repeat_customers = sum(1 for count in customer_order_count.values() if count > 1)

            # Top customers
            top_customers = sorted(
                [
                    {
                        "customer_id": cust_id,
                        "spending": spending,
                        "orders": customer_order_count.get(cust_id, 0)
                    }
                    for cust_id, spending in customer_spending.items()
                ],
                key=lambda x: x["spending"],
                reverse=True
            )[:10]

            # Customer LTV (Lifetime Value)
            total_customer_spending = sum(customer_spending.values())
            avg_ltv = total_customer_spending / len(customer_spending) if customer_spending else 0

            # Customer retention
            retention_rate = (repeat_customers / total_customers * 100) if total_customers > 0 else 0

            # Average order frequency
            avg_frequency = len(orders) / total_customers if total_customers > 0 else 0

            # Customer segments
            high_value_threshold = total_customer_spending / len(customer_spending) * 2 if customer_spending else 0
            segments = {
                "HIGH_VALUE": sum(1 for s in customer_spending.values() if s > high_value_threshold),
                "MEDIUM_VALUE": sum(1 for s in customer_spending.values() if s <= high_value_threshold and s > high_value_threshold / 2),
                "LOW_VALUE": sum(1 for s in customer_spending.values() if s <= high_value_threshold / 2),
                "INACTIVE": total_customers - len(customer_spending)
            }

            return {
                "total_customers": total_customers,
                "new_customers": new_customers,
                "repeat_customers": repeat_customers,
                "customer_retention": round(retention_rate, 2),
                "average_customer_ltv": round(avg_ltv, 2),
                "average_order_frequency": round(avg_frequency, 2),
                "top_customers": top_customers[:5],
                "customer_segments": [
                    {"segment": seg, "count": count}
                    for seg, count in segments.items()
                ],
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat()
                }
            }

        except Exception as e:
            raise Exception(f"Customer metrics error: {str(e)}")

    # ==================== DELIVERY ANALYTICS ====================

    @staticmethod
    async def get_delivery_metrics(start_date: Optional[str] = None,
                                   end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Get delivery performance metrics
        
        Returns:
        {
            "total_deliveries": int,
            "delivered": int,
            "failed": int,
            "pending": int,
            "on_time_delivery": float (percentage),
            "average_delivery_time": float (hours),
            "delivery_boys_performance": [{"id": "...", "deliveries": 50, "rating": 4.8}],
            "delivery_status_breakdown": [{"status": "DELIVERED", "count": 100}]
        }
        """
        try:
            if not end_date:
                end_date = datetime.now()
            else:
                end_date = datetime.fromisoformat(end_date)

            if not start_date:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = datetime.fromisoformat(start_date)

            # Get all deliveries in period
            deliveries = await db.delivery_statuses.find({
                "created_at": {
                    "$gte": start_date,
                    "$lte": end_date
                }
            }).to_list(None)

            total_deliveries = len(deliveries)

            # Status breakdown
            status_breakdown = {}
            on_time_count = 0
            delivery_times = []

            for delivery in deliveries:
                status = delivery.get("status", "UNKNOWN")
                status_breakdown[status] = status_breakdown.get(status, 0) + 1

                # On-time delivery calculation
                if status == "DELIVERED":
                    created = delivery.get("created_at", datetime.now())
                    delivered = delivery.get("delivered_at", datetime.now())
                    delivery_time = (delivered - created).total_seconds() / 3600  # hours

                    delivery_times.append(delivery_time)

                    # Assume on-time if delivered within 24 hours
                    if delivery_time <= 24:
                        on_time_count += 1

            # Delivery boy performance
            db_performance = {}
            deliveries_by_db = await db.delivery_boys_v2.find().to_list(None)

            for db in deliveries_by_db:
                db_id = str(db.get("_id", ""))
                db_deliveries = [d for d in deliveries if d.get("delivery_boy_id") == db_id]
                db_performance[db_id] = {
                    "name": db.get("name", "Unknown"),
                    "deliveries": len(db_deliveries),
                    "rating": db.get("rating", 0),
                    "phone": db.get("phone", "")
                }

            delivery_performance = sorted(
                [
                    {"id": db_id, **perf}
                    for db_id, perf in db_performance.items()
                ],
                key=lambda x: x["deliveries"],
                reverse=True
            )[:10]

            # Calculate metrics
            on_time_percentage = (on_time_count / total_deliveries * 100) if total_deliveries > 0 else 0
            avg_delivery_time = sum(delivery_times) / len(delivery_times) if delivery_times else 0

            delivered_count = status_breakdown.get("DELIVERED", 0)
            failed_count = status_breakdown.get("FAILED", 0) + status_breakdown.get("CANCELLED", 0)
            pending_count = status_breakdown.get("PENDING", 0) + status_breakdown.get("IN_TRANSIT", 0)

            return {
                "total_deliveries": total_deliveries,
                "delivered": delivered_count,
                "failed": failed_count,
                "pending": pending_count,
                "on_time_delivery_percentage": round(on_time_percentage, 2),
                "average_delivery_time_hours": round(avg_delivery_time, 2),
                "delivery_boys_performance": delivery_performance,
                "delivery_status_breakdown": [
                    {"status": status, "count": count}
                    for status, count in sorted(status_breakdown.items())
                ],
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat()
                }
            }

        except Exception as e:
            raise Exception(f"Delivery metrics error: {str(e)}")

    # ==================== INVENTORY ANALYTICS ====================

    @staticmethod
    async def get_inventory_insights() -> Dict[str, Any]:
        """
        Get inventory management insights
        
        Returns:
        {
            "total_products": int,
            "low_stock_items": [{"product_id": "...", "stock": 5}],
            "bestsellers": [{"product": "Milk", "units_sold": 1000}],
            "slow_movers": [{"product": "...", "units_sold": 10}],
            "stock_value": float,
            "stockout_risk": [{"product": "...", "days_to_stockout": 2}]
        }
        """
        try:
            # Get all products
            products = await db.products.find().to_list(None)
            total_products = len(products)

            # Get product sales (units sold)
            orders = await db.orders.find().to_list(None)
            product_sales = {}
            low_stock_threshold = 10

            for order in orders:
                items = order.get("items", [])
                for item in items:
                    product_name = item.get("product_name", "Unknown")
                    quantity = item.get("quantity", 0)

                    if product_name not in product_sales:
                        product_sales[product_name] = 0
                    product_sales[product_name] += quantity

            # Bestsellers (top 5)
            bestsellers = sorted(
                [
                    {"product": product, "units_sold": quantity}
                    for product, quantity in product_sales.items()
                ],
                key=lambda x: x["units_sold"],
                reverse=True
            )[:5]

            # Slow movers
            slow_movers = sorted(
                [
                    {"product": product, "units_sold": quantity}
                    for product, quantity in product_sales.items()
                ],
                key=lambda x: x["units_sold"]
            )[:5]

            # Low stock items
            low_stock_items = []
            for product in products:
                stock = product.get("stock", 0)
                if stock < low_stock_threshold:
                    low_stock_items.append({
                        "product_id": str(product.get("_id", "")),
                        "product_name": product.get("name", "Unknown"),
                        "stock": stock,
                        "reorder_level": low_stock_threshold
                    })

            # Stock value
            stock_value = sum(
                float(product.get("price", 0)) * product.get("stock", 0)
                for product in products
            )

            # Stockout risk (products running out in <7 days based on sales velocity)
            stockout_risk = []
            for product_name, daily_sales in product_sales.items():
                avg_daily_sales = daily_sales / 30  # Assuming 30 days data
                if avg_daily_sales > 0:
                    matching_products = [p for p in products if p.get("name") == product_name]
                    if matching_products:
                        product = matching_products[0]
                        stock = product.get("stock", 0)
                        days_to_stockout = stock / avg_daily_sales if avg_daily_sales > 0 else float('inf')

                        if days_to_stockout < 7:
                            stockout_risk.append({
                                "product": product_name,
                                "current_stock": stock,
                                "daily_sales": round(avg_daily_sales, 1),
                                "days_to_stockout": round(days_to_stockout, 1)
                            })

            return {
                "total_products": total_products,
                "low_stock_items": low_stock_items,
                "bestsellers": bestsellers,
                "slow_movers": slow_movers,
                "total_stock_value": round(stock_value, 2),
                "stockout_risk": sorted(
                    stockout_risk,
                    key=lambda x: x["days_to_stockout"]
                )
            }

        except Exception as e:
            raise Exception(f"Inventory insights error: {str(e)}")

    # ==================== EXPORT FUNCTIONALITY ====================

    @staticmethod
    def generate_csv_export(data: Dict[str, Any], report_type: str) -> str:
        """Generate CSV format export"""
        import csv
        from io import StringIO

        output = StringIO()
        
        if report_type == "revenue":
            writer = csv.DictWriter(
                output,
                fieldnames=["Date", "Revenue", "Orders", "AOV"]
            )
            writer.writeheader()
            
            for i, daily_rev in enumerate(data.get("daily_revenue", [])):
                daily_ord = data.get("daily_orders", [])[i] if i < len(data.get("daily_orders", [])) else {}
                writer.writerow({
                    "Date": daily_rev.get("date", ""),
                    "Revenue": daily_rev.get("amount", 0),
                    "Orders": daily_ord.get("count", 0),
                    "AOV": daily_rev.get("amount", 0) / daily_ord.get("count", 1) if daily_ord.get("count", 0) > 0 else 0
                })

        elif report_type == "customers":
            writer = csv.DictWriter(
                output,
                fieldnames=["Metric", "Value"]
            )
            writer.writeheader()
            writer.writerow({"Metric": "Total Customers", "Value": data.get("total_customers", 0)})
            writer.writerow({"Metric": "New Customers", "Value": data.get("new_customers", 0)})
            writer.writerow({"Metric": "Repeat Rate %", "Value": data.get("customer_retention", 0)})
            writer.writerow({"Metric": "Average LTV", "Value": data.get("average_customer_ltv", 0)})

        return output.getvalue()

    @staticmethod
    def generate_json_export(data: Dict[str, Any]) -> str:
        """Generate JSON format export"""
        return json.dumps(data, indent=2, default=str)

    @staticmethod
    def generate_excel_export(data: Dict[str, Any], report_type: str) -> bytes:
        """Generate Excel format export"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = report_type.capitalize()

            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            if report_type == "revenue":
                headers = ["Date", "Revenue", "Orders", "AOV"]
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                for i, daily_rev in enumerate(data.get("daily_revenue", [])):
                    daily_ord = data.get("daily_orders", [])[i] if i < len(data.get("daily_orders", [])) else {}
                    aov = daily_rev.get("amount", 0) / daily_ord.get("count", 1) if daily_ord.get("count", 0) > 0 else 0
                    ws.append([
                        daily_rev.get("date", ""),
                        daily_rev.get("amount", 0),
                        daily_ord.get("count", 0),
                        aov
                    ])

            # Write to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            raise Exception("openpyxl not installed for Excel export")

    @staticmethod
    def generate_pdf_export(data: Dict[str, Any], report_type: str) -> bytes:
        """Generate PDF format export"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from io import BytesIO

            output = BytesIO()
            c = canvas.Canvas(output, pagesize=letter)
            
            # Title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, 750, f"{report_type.capitalize()} Report")
            
            # Metadata
            c.setFont("Helvetica", 10)
            y = 720
            for key, value in data.items():
                if not isinstance(value, list):
                    c.drawString(50, y, f"{key}: {value}")
                    y -= 20

            c.save()
            return output.getvalue()

        except ImportError:
            raise Exception("reportlab not installed for PDF export")

    @staticmethod
    def generate_html_export(data: Dict[str, Any], report_type: str) -> str:
        """Generate HTML format export"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{report_type.capitalize()} Report</title>
            <style>
                body {{ font-family: Arial; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4472C4; color: white; }}
                .metric {{ font-weight: bold; }}
            </style>
        </head>
        <body>
            <h1>{report_type.capitalize()} Report</h1>
            <table>
                <tr>
        """

        if report_type == "revenue":
            html += "<th>Date</th><th>Revenue</th><th>Orders</th></tr>"
            for daily in data.get("daily_revenue", []):
                html += f"<tr><td>{daily.get('date')}</td><td>₹{daily.get('amount')}</td></tr>"

        html += """
            </table>
        </body>
        </html>
        """
        return html
